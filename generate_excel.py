import openpyxl
from openpyxl import Workbook

# Criar um novo workbook
wb = Workbook()
ws = wb.active
ws.title = "LoginData"

# Adicionar headers
ws['A1'] = 'email'
ws['B1'] = 'senha'
ws['C1'] = 'nome'
ws['D1'] = 'perfil'

# Adicionar dados de teste
ws['A2'] = 'user1@example.com'
ws['B2'] = 'senha123'
ws['C2'] = 'Usuário Teste 1'
ws['D2'] = 'admin'

ws['A3'] = 'user2@example.com'
ws['B3'] = 'senha456'
ws['C3'] = 'Usuário Teste 2'
ws['D3'] = 'usuario'

ws['A4'] = 'invalid@example.com'
ws['B4'] = 'senhaerrada'
ws['C4'] = 'Usuário Inválido'
ws['D4'] = 'usuario'

# Salvar o arquivo
wb.save('src/test/resources/data/LoginData.xlsx')
print('Arquivo LoginData.xlsx criado com sucesso!')
